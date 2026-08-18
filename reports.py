"""
reports.py — خروجی اکسل «بورس‌نگار»
Excel export for the market-gainer tables, built with openpyxl (same dependency
the original insert scripts used).
"""
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import db

_HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
_HEAD_FONT = Font(bold=True, color="FFFFFF", name="Tahoma", size=10)
_CELL_FONT = Font(name="Tahoma", size=10)
_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center")


def gainer_workbook(rows, title, as_of):
    wb = Workbook()
    ws = wb.active
    ws.title = "بازدهی"
    ws.sheet_view.rightToLeft = True

    headers = ["نماد", "نام", "گروه", "قیمت پایانی"] + [p["label"] for p in db.PERIODS]
    ws.append(headers)
    for c in ws[1]:
        c.fill, c.font, c.alignment, c.border = _HEAD_FILL, _HEAD_FONT, _CENTER, _BORDER

    for r in rows:
        line = [r["ticker"], r["name"], r.get("market") or r.get("type") or "",
                round(r["latest"]) if r["latest"] is not None else None]
        for p in db.PERIODS:
            v = r.get(p["key"])
            line.append(round(v, 2) if v is not None else None)
        ws.append(line)
        for c in ws[ws.max_row]:
            c.font, c.border = _CELL_FONT, _BORDER
            c.alignment = _CENTER

    widths = [12, 34, 22, 14] + [12] * len(db.PERIODS)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"

    # a small caption row underneath
    ws.append([])
    ws.append([f"{title} — تاریخ مبنا: {as_of}"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
